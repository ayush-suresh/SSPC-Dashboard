"""
extractor.py  —  reads AuST and CP DOR .xlsm files.
Column positions are detected dynamically from headers so the extractor
works even when new products or columns are added to the file.
"""
import io
import calendar
from collections import defaultdict
import pandas as pd
from openpyxl import load_workbook

# ── Cost tables ───────────────────────────────────────────────────────────────
COSTS = {
    "BLN": [17.93, 27.08], "SMG": [18.11, 26.57],
    "CTI": [17.93, 27.08], "BMD": [17.93, 27.08],
    "MP3": [9.86,  13.23], "NFE": [17.93, 27.08],
    "BLT": [17.93, 27.08],
}
DEFECT_COST_AUST = {
    "Leak": 16.43, "Leak - valve": 16.43, "Leak - bond": 16.43,
    "Scratch": 16.43, "Dirty": 16.43, "Wire at Tip": 11.08,
    "Wire at Hub": 13.38, "Flash ID": 11.08, "Destroyed Tip": 11.08,
    "OD Bulge": 16.43, "Gap during Reflow": 9.21, "Void at Hub Joint": 13.38,
    "OD Flash at Hub Joint": 11.08, "OD Tip Flash": 11.08,
    "Side Exposed Wire": 16.43, "Bubble": 9.21,
    "Fiber/ Embedded Particulate": 13.38, "Dent": 13.38, "Skive": 9.21,
    "Kink": 16.43, "Burn": 16.43, "Damaged/ Melted Hub": 16.43,
    "Bad Cut Valve Visual": 16.43, "Unknown Reflow Time": 9.21,
    "Cut Short": 16.43, "Pin Holes in Heat Shrink": 9.21,
    "Failed Shape (Handle Wrong Direction/ doesn't conform to template)": 16.43,
    "Extrusions in Wrong Order": 9.21, "Tip Bleed": 11.08,
    "Irregular Braid": 1.82, "Hole at Tip": 11.08, "Marker": 16.43,
    "Wrong Valve Position": 16.43, "Flash OD": 11.08,
    "Stress Marks (Hub)": 16.43, "Destructive Test": 16.43, "Other": 16.43,
}
DEFECT_COST_CP = {k: 22.08 for k in DEFECT_COST_AUST}
DEFECT_COST_CP.update({
    "Leak": 22.51, "Leak - valve": 22.51, "Leak - bond": 22.51,
    "Unknown Reflow Time": 21.94, "Extrusions in Wrong Order": 21.94,
    "Destructive Test": 22.08,
})
DEFECT_COST_MAP = {"AuST": DEFECT_COST_AUST, "CenterPoint": DEFECT_COST_CP}

# ── AuST scrap reason order (fixed order within each product block) ────────────
AUST_REASONS = [
    "Bad Liner", "Irregular Braid", "Open Braid", "Unknow Reflow Time",
    "Scratch", "Dirty", "Fiber/Emb. Part.", "Skive", "Kink", "Burn",
    "Destroyed Tip", "Tip Bleed", "Valve Orientation", "Cut Short",
    "OD Tip Flash", "Wire at Tip", "Damaged/Melt Hub", "Pin Holes at HS",
    "Hole at Tip", "Marker", "Flash OD", "Leak - valve", "Leak - bond",
    "Destructive Test", "Gap", "Layup Wrong Extrusion Order",
    "Irregular Liner at Tip", "Stress Marks (Hub)", "other",
]
AUST_MAP = {
    "Bad Liner": None, "Irregular Braid": "Irregular Braid",
    "Open Braid": "Other", "Unknow Reflow Time": "Unknown Reflow Time",
    "Scratch": "Scratch", "Dirty": "Dirty",
    "Fiber/Emb. Part.": "Fiber/ Embedded Particulate",
    "Skive": "Skive", "Kink": "Kink", "Burn": "Burn",
    "Destroyed Tip": "Destroyed Tip", "Tip Bleed": "Tip Bleed",
    "Valve Orientation": "Wrong Valve Position", "Cut Short": "Cut Short",
    "OD Tip Flash": None, "Wire at Tip": "Wire at Tip",
    "Damaged/Melt Hub": "Damaged/ Melted Hub",
    "Pin Holes at HS": "Pin Holes in Heat Shrink",
    "Hole at Tip": "Hole at Tip", "Marker": "Marker",
    "Flash OD": "Flash OD", "Leak - valve": "Leak - valve",
    "Leak - bond": "Leak - bond", "Destructive Test": "Destructive Test",
    "Gap": "Gap during Reflow",
    "Layup Wrong Extrusion Order": "Extrusions in Wrong Order",
    "Irregular Liner at Tip": "Irregular Liner at Tip",
    "Stress Marks (Hub)": "Stress Marks (Hub)", "other": "Other",
}

# ── CP scrap reason order ─────────────────────────────────────────────────────
CP_REASONS = [
    "Leak - Bond", "Leak - Valve", "Scratch", "Dirty", "Wire At Tip",
    "Flash ID", "Destroyed Tip", "Od Tip Flash", "Fiber/Embedded Particulate",
    "Skive", "Kink", "Burn", "Damage/Melted Hub", "Unknown Reflow Time",
    "Cut Short", "Pin Holes in Heat Shrink",
    "Failed Shape (Handle Wrong Direction/doest conform to template)",
    "Tip Bleed", "Irregular Bread", "Hole at Tip", "Marker",
    "Bad Liner (Splotchy coating)", "Valve Deformed/Dammage/dirty", "other",
]
CP_MAP = {
    "Leak - Bond": "Leak - bond", "Leak - Valve": "Leak - valve",
    "Scratch": "Scratch", "Dirty": "Dirty", "Wire At Tip": "Wire at Tip",
    "Flash ID": None, "Destroyed Tip": "Destroyed Tip", "Od Tip Flash": None,
    "Fiber/Embedded Particulate": "Fiber/ Embedded Particulate",
    "Skive": "Skive", "Kink": "Kink", "Burn": "Burn",
    "Damage/Melted Hub": "Damaged/ Melted Hub",
    "Unknown Reflow Time": "Unknown Reflow Time",
    "Cut Short": "Cut Short",
    "Pin Holes in Heat Shrink": "Pin Holes in Heat Shrink",
    "Failed Shape (Handle Wrong Direction/doest conform to template)":
        "Failed Shape (Handle Wrong Direction/ doesn't conform to template)",
    "Tip Bleed": "Tip Bleed", "Irregular Bread": "Irregular Braid",
    "Hole at Tip": "Hole at Tip", "Marker": "Marker",
    "Bad Liner (Splotchy coating)": None,
    "Valve Deformed/Dammage/dirty": "Bad Cut Valve Visual",
    "other": "Other",
}

ALL_DEFECT_COLS = [
    "Leak", "Leak - valve", "Leak - bond", "Scratch", "Dirty",
    "Wire at Tip", "Wire at Hub", "Flash ID", "Destroyed Tip", "OD Bulge",
    "Gap during Reflow", "Void at Hub Joint", "OD Flash at Hub Joint",
    "OD Tip Flash", "Side Exposed Wire", "Bubble",
    "Fiber/ Embedded Particulate", "Dent", "Skive", "Kink", "Burn",
    "Damaged/ Melted Hub", "Bad Cut Valve Visual", "Unknown Reflow Time",
    "Cut Short", "Pin Holes in Heat Shrink",
    "Failed Shape (Handle Wrong Direction/ doesn't conform to template)",
    "Extrusions in Wrong Order", "Tip Bleed", "Irregular Braid",
    "Hole at Tip", "Marker", "Wrong Valve Position", "Flash OD",
    "Stress Marks (Hub)", "Irregular Liner at Tip", "Destructive Test", "Other",
]
EXCLUDE_FROM_TOP10 = {
    "Other", "Unknown Issue Samples in Retains",
    "Bad Liner (splotchy coating, etc)", "Leak",
}


# ── Hard-coded historical COPQ (May 2023 – Dec 2025) ─────────────────────────
HISTORICAL_COPQ = [
    {'year':2023,'month':5,'aust_good':69878.025,'aust_scrap':12569.644,'cp_good':0.0,'cp_scrap':0.0,'total_good':69878.025,'total_scrap':12569.644,'costed_yield':0.847544,'copq_per_part':None,'leak_aust':397.0,'leak_cp':0.0,'good_aust':3858.0,'good_cp':0.0,'leak_rate_aust':0.093302,'leak_rate_cp':None,'cumul_leak':None,'smg_yield':None,'bln_yield':None,'cti_yield':None,'bmd_yield':None,'mp3_yield':None},
    {'year':2023,'month':6,'aust_good':72848.475,'aust_scrap':17242.1815,'cp_good':0.0,'cp_scrap':0.0,'total_good':72848.475,'total_scrap':17242.1815,'costed_yield':0.808613,'copq_per_part':None,'leak_aust':486.0,'leak_cp':0.0,'good_aust':4022.0,'good_cp':0.0,'leak_rate_aust':0.107808,'leak_rate_cp':None,'cumul_leak':None,'smg_yield':None,'bln_yield':None,'cti_yield':None,'bmd_yield':None,'mp3_yield':None},
    {'year':2023,'month':7,'aust_good':77260.785,'aust_scrap':11199.5975,'cp_good':0.0,'cp_scrap':0.0,'total_good':77260.785,'total_scrap':11199.5975,'costed_yield':0.873394,'copq_per_part':None,'leak_aust':95.0,'leak_cp':0.0,'good_aust':4283.0,'good_cp':0.0,'leak_rate_aust':0.021699,'leak_rate_cp':None,'cumul_leak':None,'smg_yield':None,'bln_yield':None,'cti_yield':None,'bmd_yield':None,'mp3_yield':None},
    {'year':2023,'month':8,'aust_good':120301.7865,'aust_scrap':5550.3075,'cp_good':0.0,'cp_scrap':0.0,'total_good':120301.7865,'total_scrap':5550.3075,'costed_yield':0.955898,'copq_per_part':None,'leak_aust':112.0,'leak_cp':0.0,'good_aust':6661.0,'good_cp':0.0,'leak_rate_aust':0.016536,'leak_rate_cp':None,'cumul_leak':None,'smg_yield':None,'bln_yield':None,'cti_yield':None,'bmd_yield':None,'mp3_yield':None},
    {'year':2023,'month':9,'aust_good':119069.5695,'aust_scrap':4775.414,'cp_good':0.0,'cp_scrap':0.0,'total_good':119069.5695,'total_scrap':4775.414,'costed_yield':0.96144,'copq_per_part':None,'leak_aust':112.0,'leak_cp':0.0,'good_aust':6599.0,'good_cp':0.0,'leak_rate_aust':0.016689,'leak_rate_cp':None,'cumul_leak':None,'smg_yield':None,'bln_yield':None,'cti_yield':None,'bmd_yield':None,'mp3_yield':None},
    {'year':2023,'month':10,'aust_good':122125.6155,'aust_scrap':6884.426,'cp_good':0.0,'cp_scrap':0.0,'total_good':122125.6155,'total_scrap':6884.426,'costed_yield':0.946637,'copq_per_part':None,'leak_aust':202.0,'leak_cp':0.0,'good_aust':6779.0,'good_cp':0.0,'leak_rate_aust':0.028936,'leak_rate_cp':None,'cumul_leak':None,'smg_yield':None,'bln_yield':None,'cti_yield':None,'bmd_yield':None,'mp3_yield':None},
    {'year':2023,'month':11,'aust_good':171254.895,'aust_scrap':5503.211,'cp_good':0.0,'cp_scrap':0.0,'total_good':171254.895,'total_scrap':5503.211,'costed_yield':0.968866,'copq_per_part':None,'leak_aust':136.0,'leak_cp':0.0,'good_aust':9480.0,'good_cp':0.0,'leak_rate_aust':0.014143,'leak_rate_cp':None,'cumul_leak':None,'smg_yield':None,'bln_yield':None,'cti_yield':None,'bmd_yield':None,'mp3_yield':None},
    {'year':2023,'month':12,'aust_good':133783.3245,'aust_scrap':4967.03,'cp_good':0.0,'cp_scrap':0.0,'total_good':133783.3245,'total_scrap':4967.03,'costed_yield':0.964202,'copq_per_part':None,'leak_aust':93.0,'leak_cp':0.0,'good_aust':7432.0,'good_cp':0.0,'leak_rate_aust':0.012359,'leak_rate_cp':None,'cumul_leak':None,'smg_yield':None,'bln_yield':None,'cti_yield':None,'bmd_yield':None,'mp3_yield':None},
    {'year':2024,'month':1,'aust_good':183145.8405,'aust_scrap':8731.5785,'cp_good':216179.7475,'cp_scrap':32171.403,'total_good':399325.588,'total_scrap':40902.9815,'costed_yield':0.907087,'copq_per_part':5.026789,'leak_aust':116.0,'leak_cp':958.0,'good_aust':10196.0,'good_cp':8137.0,'leak_rate_aust':0.011249,'leak_rate_cp':0.105333,'cumul_leak':0.118087,'smg_yield':None,'bln_yield':None,'cti_yield':None,'bmd_yield':None,'mp3_yield':None},
    {'year':2024,'month':2,'aust_good':187061.2275,'aust_scrap':9102.6645,'cp_good':193039.455,'cp_scrap':30643.2925,'total_good':380100.6825,'total_scrap':39745.957,'costed_yield':0.905332,'copq_per_part':5.470129,'leak_aust':167.0,'leak_cp':693.0,'good_aust':10385.0,'good_cp':7266.0,'leak_rate_aust':0.015826,'leak_rate_cp':0.087071,'cumul_leak':0.108054,'smg_yield':None,'bln_yield':None,'cti_yield':None,'bmd_yield':None,'mp3_yield':None},
    {'year':2024,'month':3,'aust_good':212036.748,'aust_scrap':6241.5035,'cp_good':285706.895,'cp_scrap':28370.222,'total_good':497743.643,'total_scrap':34611.7255,'costed_yield':0.934984,'copq_per_part':3.218498,'leak_aust':192.0,'leak_cp':791.0,'good_aust':11800.0,'good_cp':10754.0,'leak_rate_aust':0.016011,'leak_rate_cp':0.068515,'cumul_leak':0.085145,'smg_yield':None,'bln_yield':None,'cti_yield':None,'bmd_yield':None,'mp3_yield':None},
    {'year':2024,'month':4,'aust_good':207541.11,'aust_scrap':6697.449,'cp_good':258661.18,'cp_scrap':33641.066,'total_good':466202.29,'total_scrap':40338.515,'costed_yield':0.920365,'copq_per_part':4.143233,'leak_aust':120.0,'leak_cp':886.0,'good_aust':11516.0,'good_cp':9736.0,'leak_rate_aust':0.010313,'leak_rate_cp':0.083412,'cumul_leak':0.094709,'smg_yield':None,'bln_yield':None,'cti_yield':None,'bmd_yield':None,'mp3_yield':None},
    {'year':2024,'month':5,'aust_good':164772.3315,'aust_scrap':4379.2305,'cp_good':176217.88,'cp_scrap':19567.1775,'total_good':340990.2115,'total_scrap':23946.408,'costed_yield':0.934382,'copq_per_part':3.623851,'leak_aust':120.0,'leak_cp':574.0,'good_aust':9136.0,'good_cp':6608.0,'leak_rate_aust':0.012965,'leak_rate_cp':0.079922,'cumul_leak':0.09663,'smg_yield':None,'bln_yield':None,'cti_yield':None,'bmd_yield':None,'mp3_yield':None},
    {'year':2024,'month':6,'aust_good':170471.112,'aust_scrap':6518.598,'cp_good':137576.7855,'cp_scrap':22682.7685,'total_good':308047.8975,'total_scrap':29201.3665,'costed_yield':0.913413,'copq_per_part':5.715672,'leak_aust':219.0,'leak_cp':687.0,'good_aust':9478.0,'good_cp':5109.0,'leak_rate_aust':0.022584,'leak_rate_cp':0.11853,'cumul_leak':0.156315,'smg_yield':None,'bln_yield':None,'cti_yield':None,'bmd_yield':None,'mp3_yield':None},
    {'year':2024,'month':7,'aust_good':154041.1845,'aust_scrap':4075.734,'cp_good':186117.3395,'cp_scrap':19321.221,'total_good':340158.524,'total_scrap':23396.955,'costed_yield':0.935644,'copq_per_part':3.362115,'leak_aust':137.0,'leak_cp':558.0,'good_aust':8540.0,'good_cp':6959.0,'leak_rate_aust':0.015789,'leak_rate_cp':0.074232,'cumul_leak':0.092457,'smg_yield':None,'bln_yield':None,'cti_yield':None,'bmd_yield':None,'mp3_yield':None},
    {'year':2024,'month':8,'aust_good':179337.921,'aust_scrap':5025.232,'cp_good':215707.371,'cp_scrap':21865.6385,'total_good':395045.292,'total_scrap':26890.8705,'costed_yield':0.936268,'copq_per_part':3.347133,'leak_aust':133.0,'leak_cp':676.0,'good_aust':9954.0,'good_cp':8034.0,'leak_rate_aust':0.013185,'leak_rate_cp':0.077612,'cumul_leak':0.092882,'smg_yield':None,'bln_yield':None,'cti_yield':None,'bmd_yield':None,'mp3_yield':None},
    {'year':2024,'month':9,'aust_good':181708.905,'aust_scrap':5118.9405,'cp_good':183448.702,'cp_scrap':24433.5255,'total_good':365157.607,'total_scrap':29552.466,'costed_yield':0.925129,'copq_per_part':4.318011,'leak_aust':125.0,'leak_cp':576.0,'good_aust':10079.0,'good_cp':6844.0,'leak_rate_aust':0.01225,'leak_rate_cp':0.077628,'cumul_leak':0.094474,'smg_yield':None,'bln_yield':None,'cti_yield':None,'bmd_yield':None,'mp3_yield':None},
    {'year':2024,'month':10,'aust_good':206949.834,'aust_scrap':7549.01,'cp_good':209430.763,'cp_scrap':31312.4255,'total_good':416380.597,'total_scrap':38861.4355,'costed_yield':0.914636,'copq_per_part':4.965683,'leak_aust':226.0,'leak_cp':998.0,'good_aust':11464.0,'good_cp':7826.0,'leak_rate_aust':0.019333,'leak_rate_cp':0.113101,'cumul_leak':0.138713,'smg_yield':None,'bln_yield':None,'cti_yield':None,'bmd_yield':None,'mp3_yield':None},
    {'year':2024,'month':11,'aust_good':137049.8325,'aust_scrap':10911.304,'cp_good':152862.839,'cp_scrap':41239.8605,'total_good':289912.6715,'total_scrap':52151.1645,'costed_yield':0.84754,'copq_per_part':9.136504,'leak_aust':442.0,'leak_cp':1181.0,'good_aust':7593.0,'good_cp':5708.0,'leak_rate_aust':0.055009,'leak_rate_cp':0.171433,'cumul_leak':0.235593,'smg_yield':None,'bln_yield':None,'cti_yield':None,'bmd_yield':None,'mp3_yield':None},
    {'year':2024,'month':12,'aust_good':151641.0105,'aust_scrap':6289.477,'cp_good':221971.299,'cp_scrap':38064.845,'total_good':373612.3095,'total_scrap':44354.322,'costed_yield':0.893881,'copq_per_part':5.391967,'leak_aust':236.0,'leak_cp':1033.0,'good_aust':8416.0,'good_cp':8226.0,'leak_rate_aust':0.027277,'leak_rate_cp':0.111567,'cumul_leak':0.137056,'smg_yield':None,'bln_yield':None,'cti_yield':None,'bmd_yield':None,'mp3_yield':None},
    {'year':2025,'month':1,'aust_good':2439.024,'aust_scrap':0.0,'cp_good':18331.575,'cp_scrap':1249.6885,'total_good':20770.599,'total_scrap':1249.6885,'costed_yield':0.943248,'copq_per_part':1.811143,'leak_aust':0.0,'leak_cp':51.0,'good_aust':136.0,'good_cp':690.0,'leak_rate_aust':0.0,'leak_rate_cp':0.068826,'cumul_leak':0.068826,'smg_yield':None,'bln_yield':None,'cti_yield':None,'bmd_yield':None,'mp3_yield':None},
    {'year':2025,'month':2,'aust_good':282227.358,'aust_scrap':14692.154,'cp_good':222549.0365,'cp_scrap':30758.9055,'total_good':504776.3945,'total_scrap':45451.0595,'costed_yield':0.917396,'copq_per_part':5.503216,'leak_aust':462.0,'leak_cp':376.0,'good_aust':15737.0,'good_cp':8259.0,'leak_rate_aust':0.02852,'leak_rate_cp':0.043544,'cumul_leak':0.097047,'smg_yield':0.922083,'bln_yield':0.939271,'cti_yield':0.879928,'bmd_yield':0.919872,'mp3_yield':None},
    {'year':2025,'month':3,'aust_good':225179.304,'aust_scrap':12340.822,'cp_good':217995.376,'cp_scrap':33833.4845,'total_good':443174.68,'total_scrap':46174.3065,'costed_yield':0.905641,'copq_per_part':5.720306,'leak_aust':398.0,'leak_cp':328.0,'good_aust':12556.0,'good_cp':8072.0,'leak_rate_aust':0.030724,'leak_rate_cp':0.039048,'cumul_leak':0.086429,'smg_yield':0.939711,'bln_yield':0.926716,'cti_yield':0.875943,'bmd_yield':None,'mp3_yield':None},
    {'year':2025,'month':4,'aust_good':229725.846,'aust_scrap':23772.7455,'cp_good':260884.384,'cp_scrap':25682.582,'total_good':490610.23,'total_scrap':49455.3275,'costed_yield':0.908427,'copq_per_part':5.125967,'leak_aust':1024.0,'leak_cp':205.0,'good_aust':12800.0,'good_cp':9648.0,'leak_rate_aust':0.074074,'leak_rate_cp':0.020806,'cumul_leak':0.124734,'smg_yield':0.901981,'bln_yield':0.951197,'cti_yield':0.86017,'bmd_yield':None,'mp3_yield':None},
    {'year':2025,'month':5,'aust_good':256620.7875,'aust_scrap':21334.9435,'cp_good':288042.555,'cp_scrap':31529.7975,'total_good':544663.3425,'total_scrap':52864.741,'costed_yield':0.911528,'copq_per_part':4.954521,'leak_aust':812.0,'leak_cp':520.0,'good_aust':14298.0,'good_cp':10670.0,'leak_rate_aust':0.053739,'leak_rate_cp':0.04647,'cumul_leak':0.119035,'smg_yield':0.925085,'bln_yield':0.897546,'cti_yield':0.923308,'bmd_yield':0.915521,'mp3_yield':None},
    {'year':2025,'month':6,'aust_good':289029.7935,'aust_scrap':28770.019,'cp_good':263983.618,'cp_scrap':37736.955,'total_good':553013.4115,'total_scrap':66506.974,'costed_yield':0.892648,'copq_per_part':6.806568,'leak_aust':1357.0,'leak_cp':622.0,'good_aust':16081.0,'good_cp':9771.0,'leak_rate_aust':0.077819,'leak_rate_cp':0.059848,'cumul_leak':0.190417,'smg_yield':0.807686,'bln_yield':0.915056,'cti_yield':0.892719,'bmd_yield':0.927588,'mp3_yield':None},
    {'year':2025,'month':7,'aust_good':282774.4185,'aust_scrap':24865.424,'cp_good':321813.2325,'cp_scrap':44314.8415,'total_good':604587.651,'total_scrap':69180.2655,'costed_yield':0.897323,'copq_per_part':5.805662,'leak_aust':751.0,'leak_cp':849.0,'good_aust':15711.0,'good_cp':11916.0,'leak_rate_aust':0.04562,'leak_rate_cp':0.06651,'cumul_leak':0.125343,'smg_yield':0.857778,'bln_yield':0.907173,'cti_yield':0.935242,'bmd_yield':0.886802,'mp3_yield':None},
    {'year':2025,'month':8,'aust_good':318497.5605,'aust_scrap':37209.2715,'cp_good':401369.061,'cp_scrap':48597.6955,'total_good':719866.6215,'total_scrap':85806.967,'costed_yield':0.893497,'copq_per_part':5.758084,'leak_aust':1335.0,'leak_cp':1132.0,'good_aust':17712.0,'good_cp':14902.0,'leak_rate_aust':0.07009,'leak_rate_cp':0.0706,'cumul_leak':0.153861,'smg_yield':0.865847,'bln_yield':0.913841,'cti_yield':0.895368,'bmd_yield':0.895812,'mp3_yield':0.827622},
    {'year':2025,'month':9,'aust_good':326396.196,'aust_scrap':43652.402,'cp_good':329995.122,'cp_scrap':41347.1995,'total_good':656391.318,'total_scrap':84999.6015,'costed_yield':0.885351,'copq_per_part':6.964897,'leak_aust':1586.0,'leak_cp':703.0,'good_aust':18152.0,'good_cp':12204.0,'leak_rate_aust':0.080353,'leak_rate_cp':0.054467,'cumul_leak':0.177346,'smg_yield':0.860637,'bln_yield':0.899906,'cti_yield':0.883093,'bmd_yield':0.850129,'mp3_yield':0.880115},
    {'year':2025,'month':10,'aust_good':426980.0535,'aust_scrap':44882.9845,'cp_good':379110.297,'cp_scrap':150930.4685,'total_good':806090.3505,'total_scrap':195813.453,'costed_yield':0.804559,'copq_per_part':13.927979,'leak_aust':1123.0,'leak_cp':674.0,'good_aust':23745.0,'good_cp':14059.0,'leak_rate_aust':0.045158,'leak_rate_cp':0.045748,'cumul_leak':0.121971,'smg_yield':0.708918,'bln_yield':0.792093,'cti_yield':0.897185,'bmd_yield':0.870301,'mp3_yield':None},
    {'year':2025,'month':11,'aust_good':336541.4745,'aust_scrap':32166.1395,'cp_good':531966.334,'cp_scrap':95346.64,'total_good':868507.8085,'total_scrap':127512.7795,'costed_yield':0.871978,'copq_per_part':6.461251,'leak_aust':671.0,'leak_cp':762.0,'good_aust':18720.0,'good_cp':19735.0,'leak_rate_aust':0.034604,'leak_rate_cp':0.037176,'cumul_leak':0.069913,'smg_yield':0.835178,'bln_yield':0.88826,'cti_yield':0.863681,'bmd_yield':0.869534,'mp3_yield':0.963261},
    {'year':2025,'month':12,'aust_good':425822.67,'aust_scrap':34901.8055,'cp_good':572658.1645,'cp_scrap':76704.5135,'total_good':998480.8345,'total_scrap':111606.319,'costed_yield':0.899462,'copq_per_part':5.25503,'leak_aust':1078.0,'leak_cp':735.0,'good_aust':23692.0,'good_cp':21238.0,'leak_rate_aust':0.04352,'leak_rate_cp':0.03345,'cumul_leak':0.08251,'smg_yield':0.861919,'bln_yield':0.91276,'cti_yield':0.924898,'bmd_yield':0.843993,'mp3_yield':0.8988},
]


def _sf(v):
    if v is None: return 0.0
    try:
        f = float(v)
        return 0.0 if f != f else f
    except: return 0.0


# ══════════════════════════════════════════════════════════════════════════════
# DYNAMIC COLUMN DETECTION
# ══════════════════════════════════════════════════════════════════════════════
def _detect_aust_cols(ws):
    """Read headers from the worksheet to find all column positions."""
    row1 = {}; row2 = {}; row3 = {}
    for c in range(1, ws.max_column + 1):
        v1 = ws.cell(1, c).value
        v2 = ws.cell(2, c).value
        v3 = ws.cell(3, c).value
        if v1 and v1 not in row1: row1[v1] = c
        if v2 and v2 not in row2: row2[v2] = c
        if v3 and v3 not in row3: row3[v3] = c

    qty_col   = row2.get("Service", 13)
    lot_col   = row2.get("LOT", 22)
    shift_col = row2.get("Shift", 7)
    date_col  = 2

    # Service cols per product (row 3)
    svc_cols = {}
    for prod in ["SMG","BLN","CTI","BMD","MP3-S","MP3-M","NFE","BLT"]:
        if prod in row3: svc_cols[prod] = row3[prod]

    # Product group header cols (row 1)
    prod_headers = {}
    for prod in ["SMG","BLN","CTI","BMD","MP3-S","MP3-M","NFE","BLT"]:
        if prod in row1: prod_headers[prod] = row1[prod]

    # Find scrap start col = 'Bad Liner' after each product header
    prod_sorted = sorted(prod_headers.items(), key=lambda x: x[1])
    scrap_starts = {}
    for i, (prod, start_c) in enumerate(prod_sorted):
        end_c = prod_sorted[i+1][1] if i+1 < len(prod_sorted) else ws.max_column
        for c in range(start_c, end_c):
            if ws.cell(2, c).value == "Bad Liner":
                scrap_starts[prod] = c
                break

    # People/hours/notes — search by header name, fall back to last occurrence
    people_col  = None; hours_col = None
    peoples_col = None; notes_col = None
    for c in range(ws.max_column, 0, -1):
        v = ws.cell(2, c).value
        if v == "# of People"            and not people_col:  people_col  = c
        if v == "Shift Hours"             and not hours_col:   hours_col   = c
        if v == "PEOPLES"                 and not peoples_col: peoples_col = c
        if v == "NOTES"                   and not notes_col:   notes_col   = c
        if people_col and hours_col and peoples_col and notes_col: break

    return {
        "qty_col": qty_col, "lot_col": lot_col, "date_col": date_col,
        "shift_col": shift_col, "svc_cols": svc_cols,
        "scrap_starts": scrap_starts,
        "people_col": people_col or 337, "hours_col": hours_col or 336,
        "peoples_col": peoples_col or 339, "notes_col": notes_col or 340,
    }


def _detect_cp_cols(ws):
    row1 = {}; row2 = {}; row3 = {}
    for c in range(1, ws.max_column + 1):
        v1 = ws.cell(1, c).value
        v2 = ws.cell(2, c).value
        v3 = ws.cell(3, c).value
        if v1 and v1 not in row1: row1[v1] = c
        if v2 and v2 not in row2: row2[v2] = c
        if v3 and v3 not in row3: row3[v3] = c

    qty_col   = row2.get("Service", 12)
    lot_col   = row2.get("LOT", 22)
    shift_col = row2.get("Shift", 7)
    date_col  = 2

    svc_cols = {}
    for prod in ["SMG","BLN","CTI","BMD","MP3-S","MP3-M","NFE","BLT"]:
        if prod in row3: svc_cols[prod] = row3[prod]

    prod_headers = {}
    for prod in ["SMG","BLN","CTI","BMD","MP3-S","MP3-M","NFE","BLT"]:
        if prod in row1: prod_headers[prod] = row1[prod]

    prod_sorted = sorted(prod_headers.items(), key=lambda x: x[1])
    scrap_starts = {}
    for i, (prod, start_c) in enumerate(prod_sorted):
        end_c = prod_sorted[i+1][1] if i+1 < len(prod_sorted) else ws.max_column
        for c in range(start_c, end_c):
            if ws.cell(2, c).value == "Leak - Bond":
                scrap_starts[prod] = c
                break

    people_col  = None; hours_col = None
    peoples_col = None; notes_col = None
    for c in range(ws.max_column, 0, -1):
        v = ws.cell(2, c).value
        if v == "# of People"         and not people_col:  people_col  = c
        if v == "# hours per shift"   and not hours_col:   hours_col   = c
        if v == "PEOPLES"             and not peoples_col: peoples_col = c
        if v == "NOTES"               and not notes_col:   notes_col   = c
        if people_col and hours_col and peoples_col and notes_col: break

    return {
        "qty_col": qty_col, "lot_col": lot_col, "date_col": date_col,
        "shift_col": shift_col, "svc_cols": svc_cols,
        "scrap_starts": scrap_starts,
        "people_col": people_col or 263, "hours_col": hours_col or 264,
        "peoples_col": peoples_col or 265, "notes_col": notes_col or 266,
    }


# ══════════════════════════════════════════════════════════════════════════════
# SCRAP EXTRACTION
# ══════════════════════════════════════════════════════════════════════════════
def _extract_scrap(ws, cols, reasons, mapping, entity):
    qty_col  = cols["qty_col"]
    lot_col  = cols["lot_col"]
    date_col = cols["date_col"]
    svc_cols = cols["svc_cols"]
    scrap_starts = cols["scrap_starts"]

    agg = defaultdict(lambda: {
        "qty": 0, "retains": 0, "lot": "",
        **{c: 0 for c in ALL_DEFECT_COLS}
    })

    for r in range(4, ws.max_row + 1):
        date = ws.cell(r, date_col).value
        if not hasattr(date, "month"): continue
        rt = ws.cell(r, qty_col).value
        if rt not in ("Qty", "Reject"): continue

        lot_raw = str(ws.cell(r, lot_col).value or "").strip()
        tokens  = [t for t in lot_raw.split()
                   if t.startswith(("ML","CL","ml","cl"))]
        lot_key = tokens[0] if tokens else (lot_raw.split()[0] if lot_raw else "")

        for prod, svc_c in svc_cols.items():
            svc = ws.cell(r, svc_c).value
            if not svc: continue
            try: count = int(float(svc))
            except: continue
            if count <= 0: continue

            # Normalise product code
            pc = ("MP3" if prod in ("MP3-S","MP3-M")
                  else "NFE" if prod in ("NFE","NFS")
                  else prod)
            key = ((date.year, date.month), pc, entity)

            if rt == "Qty":
                agg[key]["qty"] += count
                if lot_key and not agg[key]["lot"]:
                    agg[key]["lot"] = lot_key
                ss = scrap_starts.get(prod)
                if ss is None: continue
                for i, reason in enumerate(reasons):
                    dest = mapping.get(reason)
                    if dest is None: continue
                    v = ws.cell(r, ss + i).value
                    if v:
                        try: n = int(float(v))
                        except: n = 0
                        if n > 0: agg[key][dest] += n

            elif rt == "Reject":
                agg[key]["retains"] += count

    rows = []
    for (mk, pc, ent), data in agg.items():
        qty = data["qty"]; ret = data["retains"]
        if (qty == 0 and ret == 0
                and not any(data[c] > 0 for c in ALL_DEFECT_COLS)):
            continue
        data["Leak"] = data.get("Leak - valve", 0) + data.get("Leak - bond", 0)
        yld = round((qty - ret) / qty, 4) if qty > 0 else None
        row = {
            "year": mk[0], "month": mk[1], "product": pc, "entity": ent,
            "lot": data["lot"] or "", "lot_size": float(qty),
            "retains": float(ret), "yield": yld,
        }
        for c in ALL_DEFECT_COLS:
            row[c] = data.get(c, 0)
        rows.append(row)
    return rows


# ══════════════════════════════════════════════════════════════════════════════
# DAILY PRODUCTION
# ══════════════════════════════════════════════════════════════════════════════
def _extract_daily_production(ws_aust, ws_cp, aust_cols, cp_cols):
    rows = []

    def parse(ws, facility, cols, reasons_order):
        qty_col    = cols["qty_col"]
        date_col   = cols["date_col"]
        shift_col  = cols["shift_col"]
        svc_cols   = cols["svc_cols"]
        people_col = cols["people_col"]
        hours_col  = cols["hours_col"]
        peoples_col= cols["peoples_col"]
        notes_col  = cols["notes_col"]

        for r in range(4, ws.max_row + 1):
            date  = ws.cell(r, date_col).value
            if not hasattr(date, "month"): continue
            rt    = ws.cell(r, qty_col).value
            shift = ws.cell(r, shift_col).value
            if rt not in ("Qty","Reject"): continue
            if not shift: continue

            for prod, svc_c in svc_cols.items():
                val = _sf(ws.cell(r, svc_c).value)
                if val == 0: continue
                pc = ("MP3" if prod in ("MP3-S","MP3-M")
                      else "NFE" if prod in ("NFE","NFS")
                      else prod)
                rows.append({
                    "date": date, "year": date.year,
                    "month": date.month, "day": date.day,
                    "facility": facility, "shift": shift,
                    "product": pc, "row_type": rt, "value": val,
                    "people":  _sf(ws.cell(r, people_col).value),
                    "hours":   _sf(ws.cell(r, hours_col).value),
                    "peoples": str(ws.cell(r, peoples_col).value or ""),
                    "notes":   str(ws.cell(r, notes_col).value or ""),
                })

    parse(ws_aust, "AuST", aust_cols, AUST_REASONS)
    parse(ws_cp,   "CenterPoint", cp_cols, CP_REASONS)

    if not rows: return pd.DataFrame()
    df = pd.DataFrame(rows)
    qty_df = (df[df["row_type"] == "Qty"]
              [["date","year","month","day","facility","shift","product",
                "value","people","hours","peoples","notes"]]
              .rename(columns={"value": "qty"}))
    rej_df = (df[df["row_type"] == "Reject"]
              [["date","facility","shift","product","value"]]
              .rename(columns={"value": "rejects"}))
    merged = qty_df.merge(rej_df, on=["date","facility","shift","product"],
                          how="left")
    merged["rejects"]     = merged["rejects"].fillna(0)
    merged["good_units"]  = merged["qty"] - merged["rejects"]
    merged["reject_rate"] = merged.apply(
        lambda r: r["rejects"]/r["qty"] if r["qty"] > 0 else 0, axis=1)
    return merged.sort_values(
        ["date","facility","shift","product"]).reset_index(drop=True)


# ══════════════════════════════════════════════════════════════════════════════
# DEMAND / ATTENDANCE / ACTIONS
# ══════════════════════════════════════════════════════════════════════════════
def _extract_demand(ws):
    months = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(3, c).value
        if hasattr(v, "month"): months[c] = v
    rows = []
    for facility, row_range in [("AuST", range(16,25)),
                                  ("CenterPoint", range(29,38))]:
        for r in row_range:
            prod = ws.cell(r, 3).value
            if not prod or not isinstance(prod, str): continue
            for col, dt in months.items():
                val = _sf(ws.cell(r, col).value)
                rows.append({"year":dt.year,"month":dt.month,
                              "product":prod,"facility":facility,"plan":val})
    return pd.DataFrame(rows) if rows else pd.DataFrame()


def _extract_attendance(ws):
    dates = {}
    for c in range(2, ws.max_column + 1):
        v = ws.cell(3, c).value
        if hasattr(v, "month"): dates[c] = v
    STATUS = {"a": "On Time", "i": "Late", "r": "No Show"}
    rows = []
    for r in range(4, 15):
        team = ws.cell(r, 1).value
        if not team or not isinstance(team, str): continue
        for col, dt in dates.items():
            raw = ws.cell(r, col).value
            if raw is None: continue
            if isinstance(raw, str):
                status = STATUS.get(raw.lower().strip(), raw)
            elif isinstance(raw, (int, float)):
                status = f"{raw:.0%}" if 0 < raw <= 1 else str(int(raw))
            else: continue
            rows.append({"date": dt, "year": dt.year, "month": dt.month,
                          "day": dt.day, "team": team, "status": status})
    return pd.DataFrame(rows) if rows else pd.DataFrame()


def _extract_actions(ws):
    rows = []
    for r in range(2, ws.max_row + 1):
        date     = ws.cell(r, 1).value
        category = ws.cell(r, 2).value
        urgency  = ws.cell(r, 3).value
        desc     = ws.cell(r, 4).value
        task     = ws.cell(r, 5).value
        assigned = ws.cell(r, 6).value
        due      = ws.cell(r, 7).value
        status   = ws.cell(r, 8).value
        if date is None and category is None: continue
        if isinstance(date, str):
            try: date = pd.to_datetime(date, errors="coerce")
            except: date = None
        rows.append({
            "date": date,
            "year":  date.year  if hasattr(date, "year")  else None,
            "month": date.month if hasattr(date, "month") else None,
            "category":    str(category or "").strip(),
            "urgency":     str(urgency  or "").strip(),
            "description": str(desc     or "").strip(),
            "task":        str(task     or "").strip(),
            "assigned_to": str(assigned or "").strip(),
            "due_date":    due,
            "status":      str(status   or "Open").strip(),
        })
    return pd.DataFrame(rows) if rows else pd.DataFrame()


# ══════════════════════════════════════════════════════════════════════════════
# COPQ COMPUTATION
# ══════════════════════════════════════════════════════════════════════════════
def _compute_copq(df_scrap):
    months = sorted(df_scrap[["year","month"]].drop_duplicates()
                    .apply(tuple, axis=1).tolist())
    copq_rows = []
    for (yr, mo) in months:
        mdf  = df_scrap[(df_scrap["year"]==yr) & (df_scrap["month"]==mo)]
        aust = mdf[mdf["entity"]=="AuST"]
        cp   = mdf[mdf["entity"]=="CenterPoint"]

        def good_cost(sub, entity):
            total = 0.0
            for _, row in sub.iterrows():
                rate = COSTS.get(row["product"], COSTS["BLN"])[
                    0 if entity == "AuST" else 1]
                total += max(row["lot_size"] - row["retains"], 0) * rate
            return total

        def scrap_cost(sub, entity):
            dc = DEFECT_COST_MAP[entity]; total = 0.0
            for _, row in sub.iterrows():
                for col in ALL_DEFECT_COLS:
                    if col == "Leak": continue
                    q = _sf(row.get(col, 0))
                    if q > 0: total += q * dc.get(col, 16.43)
            return total

        def prod_yield(sub_all, prod):
            p = sub_all[sub_all["product"] == prod]
            if p.empty: return None
            pg = (good_cost(p[p["entity"]=="AuST"],       "AuST") +
                  good_cost(p[p["entity"]=="CenterPoint"], "CenterPoint"))
            pb = (scrap_cost(p[p["entity"]=="AuST"],       "AuST") +
                  scrap_cost(p[p["entity"]=="CenterPoint"], "CenterPoint"))
            return pg/(pg+pb) if (pg+pb) > 0 else None

        ag = good_cost(aust, "AuST");   cg = good_cost(cp, "CenterPoint")
        ab = scrap_cost(aust, "AuST");  cb = scrap_cost(cp, "CenterPoint")
        tgc = ag + cg; tbc = ab + cb

        good_cp   = sum(max(r["lot_size"]-r["retains"], 0)
                        for _, r in cp.iterrows())
        good_aust = sum(max(r["lot_size"]-r["retains"], 0)
                        for _, r in aust.iterrows())
        leak_aust = float(aust["Leak"].sum())
        leak_cp   = float(cp["Leak"].sum())
        lr_a = leak_aust/(good_aust+leak_aust) if (good_aust+leak_aust)>0 else 0
        lr_c = leak_cp  /(good_cp  +leak_cp)   if (good_cp  +leak_cp)  >0 else 0
        cumul= (leak_aust+leak_cp)/(good_cp+leak_cp) if (good_cp+leak_cp)>0 else 0

        copq_rows.append({
            "year":yr,"month":mo,
            "aust_good":ag,"aust_scrap":ab,"cp_good":cg,"cp_scrap":cb,
            "total_good":tgc,"total_scrap":tbc,
            "costed_yield":  tgc/(tgc+tbc) if (tgc+tbc)>0 else None,
            "copq_per_part": tbc/good_cp   if good_cp>0   else None,
            "leak_aust":leak_aust,"leak_cp":leak_cp,
            "good_aust":good_aust,"good_cp":good_cp,
            "leak_rate_aust":lr_a,"leak_rate_cp":lr_c,"cumul_leak":cumul,
            "bln_yield":prod_yield(mdf,"BLN"),"smg_yield":prod_yield(mdf,"SMG"),
            "cti_yield":prod_yield(mdf,"CTI"),"bmd_yield":prod_yield(mdf,"BMD"),
            "mp3_yield":prod_yield(mdf,"MP3"),
        })
    df_live = pd.DataFrame(copq_rows)

    # Merge with historical: historical fills all months before the live data starts
    df_hist = pd.DataFrame(HISTORICAL_COPQ)
    if df_live.empty:
        return df_hist.sort_values(["year","month"]).reset_index(drop=True)

    live_start = (df_live["year"].min(), df_live["month"].min())
    df_hist_filtered = df_hist[
        df_hist.apply(lambda r: (r["year"], r["month"]) < live_start, axis=1)
    ]
    combined = pd.concat([df_hist_filtered, df_live], ignore_index=True)
    return combined.sort_values(["year","month"]).reset_index(drop=True)


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def extract_from_files(aust_bytes, cp_bytes):
    errors = []
    try:
        wb = load_workbook(io.BytesIO(aust_bytes), data_only=True, keep_vba=True)
    except Exception as e:
        errors.append(f"Could not read file: {e}"); return {}, errors

    try:
        ws_aust = wb["AuST-SSPC DATA"]
        ws_cp   = wb["CP-SSPC DATA"]
    except KeyError as e:
        errors.append(f"Sheet not found: {e}"); return {}, errors

    # Detect columns dynamically
    aust_cols = _detect_aust_cols(ws_aust)
    cp_cols   = _detect_cp_cols(ws_cp)

    # Extract scrap
    aust_scrap = _extract_scrap(ws_aust, aust_cols, AUST_REASONS, AUST_MAP, "AuST")
    cp_scrap   = _extract_scrap(ws_cp,   cp_cols,   CP_REASONS,   CP_MAP,   "CenterPoint")
    df_scrap   = pd.DataFrame(aust_scrap + cp_scrap)

    df_copq    = _compute_copq(df_scrap) if not df_scrap.empty else pd.DataFrame()
    df_prod    = _extract_daily_production(ws_aust, ws_cp, aust_cols, cp_cols)

    df_demand  = _extract_demand(wb["Demand Data"]) \
                 if "Demand Data"  in wb.sheetnames else pd.DataFrame()
    df_att     = _extract_attendance(wb["Attendance"]) \
                 if "Attendance"   in wb.sheetnames else pd.DataFrame()
    df_actions = _extract_actions(wb["Actions"]) \
                 if "Actions"      in wb.sheetnames else pd.DataFrame()

    months = sorted(df_scrap[["year","month"]].drop_duplicates()
                    .apply(tuple, axis=1).tolist()) if not df_scrap.empty else []

    return {"scrap":df_scrap,"copq":df_copq,"prod":df_prod,"demand":df_demand,
            "att":df_att,"actions":df_actions,"months":months}, errors


# ══════════════════════════════════════════════════════════════════════════════
# ANALYTICS HELPERS
# ══════════════════════════════════════════════════════════════════════════════
def get_top_defects(df_scrap, year, month, top_n=10):
    mdf = df_scrap[(df_scrap["year"]==year) & (df_scrap["month"]==month)]
    totals = {}
    for col in ALL_DEFECT_COLS:
        if col in EXCLUDE_FROM_TOP10: continue
        v = float(mdf[col].sum())
        if v > 0: totals[col] = v
    total_all = sum(totals.values())
    return [(k, v, v/total_all if total_all>0 else 0)
            for k, v in sorted(totals.items(), key=lambda x: -x[1])[:top_n]]


def get_rolling_stats(df_copq, col, n=12):
    df  = df_copq.sort_values(["year","month"]).tail(n).copy()
    df["label"] = df.apply(
        lambda r: f"{calendar.month_abbr[int(r.month)]} {int(r.year)}", axis=1)
    vals = df[col].dropna().tolist()
    avg  = sum(vals)/len(vals) if vals else 0
    std  = pd.Series(vals).std() if len(vals) > 1 else 0
    return df, avg, std


def get_leak_trend(df_scrap, df_copq, n_months=12):
    months = sorted(df_copq[["year","month"]].apply(tuple,axis=1).tolist())[-n_months:]
    rows = []
    for (yr, mo) in months:
        r = df_copq[(df_copq["year"]==yr) & (df_copq["month"]==mo)]
        if r.empty: continue
        rows.append({
            "label":       f"{calendar.month_abbr[mo]} {yr}",
            "AuST":        float(r["leak_rate_aust"].iloc[0]),
            "CenterPoint": float(r["leak_rate_cp"].iloc[0]),
        })
    return pd.DataFrame(rows)


def get_tip_trend(df_scrap, df_copq, n_months=12):
    months = sorted(df_copq[["year","month"]].apply(tuple,axis=1).tolist())[-n_months:]
    rows = []
    for (yr, mo) in months:
        mdf  = df_scrap[(df_scrap["year"]==yr) & (df_scrap["month"]==mo)]
        aust = mdf[mdf["entity"]=="AuST"]
        cp   = mdf[mdf["entity"]=="CenterPoint"]
        at = float(aust["Destroyed Tip"].sum())
        al = float(aust["lot_size"].sum())
        ct = float(cp["Destroyed Tip"].sum())
        cl = float(cp["lot_size"].sum())
        rows.append({
            "label":       f"{calendar.month_abbr[mo]} {yr}",
            "AuST":        at/(at+al) if (at+al) > 0 else 0,
            "CenterPoint": ct/(ct+cl) if (ct+cl) > 0 else 0,
        })
    return pd.DataFrame(rows)


def get_leak_trend_by_product(df_scrap, df_copq, n_months=12):
    """Leak rate per product (valve+bond combined) over last n months."""
    months = sorted(df_copq[["year","month"]].apply(tuple,axis=1).tolist())[-n_months:]
    months = [(yr,mo) for yr,mo in months
              if not df_scrap[(df_scrap["year"]==yr)&(df_scrap["month"]==mo)].empty]
    products = ["BLN","SMG","CTI","BMD","MP3","NFE","BLT"]
    rows = []
    for (yr, mo) in months:
        mdf = df_scrap[(df_scrap["year"]==yr) & (df_scrap["month"]==mo)]
        row = {"label": f"{calendar.month_abbr[mo]} {yr}"}
        for prod in products:
            p = mdf[mdf["product"]==prod]
            if p.empty: continue
            leak = float(p["Leak"].sum())
            lot  = float(p["lot_size"].sum())
            if lot > 0: row[prod] = leak / lot
        rows.append(row)
    return pd.DataFrame(rows) if rows else pd.DataFrame()


def get_leak_valve_by_product(df_scrap, df_copq, n_months=12):
    """Leak-valve rate per product over last n months."""
    months = sorted(df_copq[["year","month"]].apply(tuple,axis=1).tolist())[-n_months:]
    months = [(yr,mo) for yr,mo in months
              if not df_scrap[(df_scrap["year"]==yr)&(df_scrap["month"]==mo)].empty]
    products = ["BLN","SMG","CTI","BMD","MP3","NFE","BLT"]
    rows = []
    for (yr, mo) in months:
        mdf = df_scrap[(df_scrap["year"]==yr) & (df_scrap["month"]==mo)]
        row = {"label": f"{calendar.month_abbr[mo]} {yr}"}
        for prod in products:
            p = mdf[mdf["product"]==prod]
            if p.empty: continue
            lv  = float(p["Leak - valve"].sum())
            lot = float(p["lot_size"].sum())
            if lot > 0: row[prod] = lv / lot
        rows.append(row)
    return pd.DataFrame(rows) if rows else pd.DataFrame()


def get_leak_bond_by_product(df_scrap, df_copq, n_months=12):
    """Leak-bond rate per product over last n months."""
    months = sorted(df_copq[["year","month"]].apply(tuple,axis=1).tolist())[-n_months:]
    months = [(yr,mo) for yr,mo in months
              if not df_scrap[(df_scrap["year"]==yr)&(df_scrap["month"]==mo)].empty]
    products = ["BLN","SMG","CTI","BMD","MP3","NFE","BLT"]
    rows = []
    for (yr, mo) in months:
        mdf = df_scrap[(df_scrap["year"]==yr) & (df_scrap["month"]==mo)]
        row = {"label": f"{calendar.month_abbr[mo]} {yr}"}
        for prod in products:
            p = mdf[mdf["product"]==prod]
            if p.empty: continue
            lb  = float(p["Leak - bond"].sum())
            lot = float(p["lot_size"].sum())
            if lot > 0: row[prod] = lb / lot
        rows.append(row)
    return pd.DataFrame(rows) if rows else pd.DataFrame()

